import io
import math
import os
from copy import copy

import openpyxl
from azure.storage.blob import BlobServiceClient
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font


def copy_row_format(ws, source_row, target_row):
    """
    行のフォーマット（罫線、スタイルなど）を完全にコピーする。
    値はコピーせず、フォーマットのみをコピーする。

    Args:
        ws: ワークシート
        source_row: コピー元の行番号
        target_row: コピー先の行番号
    """
    # セルのスタイルをコピー
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)
        
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
    
    # 行の高さをコピー
    if ws.row_dimensions[source_row].height:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def copy_merged_cells(ws, source_row, target_row):
    """
    結合セルの設定をコピーする。

    Args:
        ws: ワークシート
        source_row: コピー元の行番号
        target_row: コピー先の行番号
    """
    merged_cells_to_add = []
    for merged_range in list(ws.merged_cells.ranges):
        # コピー元の行の結合セルを探す
        if merged_range.min_row == source_row and merged_range.max_row == source_row:
            # 同じ列範囲でターゲット行に結合セルを追加
            new_range = f"{merged_range.coord}".replace(
                f"{source_row}", f"{target_row}"
            )
            merged_cells_to_add.append(new_range)
    
    for new_range in merged_cells_to_add:
        ws.merge_cells(new_range)


def ensure_label_row_merged_cells(ws, label_row, label_type):
    """
    ラベル行の結合セルを確保・再設定する。
    行挿入後にラベル行のフォーマットを保護するために使用。

    Args:
        ws: ワークシート
        label_row: ラベル行番号
        label_type: ラベルのタイプ（'decisions', 'homework', 'next_meeting', 'minutes'）
    """
    # ラベル行の既存の結合セルを解除（重複を防ぐ）
    merged_cells_to_remove = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row == label_row and merged_range.max_row == label_row:
            merged_cells_to_remove.append(merged_range)
    
    for merged_range in merged_cells_to_remove:
        ws.unmerge_cells(str(merged_range))
    
    # ラベルタイプに応じて結合セルを再設定
    if label_type == 'decisions':
        # 決定事項ラベル: B:AH
        ws.merge_cells(f'B{label_row}:AH{label_row}')
    elif label_type == 'homework':
        # 宿題事項ラベル: B:AD（宿題事項）、AE:AF（担当）、AG:AH（期限）
        ws.merge_cells(f'B{label_row}:AD{label_row}')
        ws.merge_cells(f'AE{label_row}:AF{label_row}')
        ws.merge_cells(f'AG{label_row}:AH{label_row}')
    elif label_type == 'next_meeting':
        # 次回予定ラベル: B:AH
        ws.merge_cells(f'B{label_row}:AH{label_row}')
    elif label_type == 'minutes':
        # 会議内容ラベル: B:AH
        ws.merge_cells(f'B{label_row}:AH{label_row}')


def get_merged_cell_parent(ws, cell):
    """
    結合セルの場合、親セル（左上のセル）を返す。
    通常のセルの場合はそのまま返す。

    Args:
        ws: ワークシート
        cell: 対象のセルオブジェクト

    Returns:
        親セルまたは通常のセルオブジェクト
    """
    if isinstance(cell, MergedCell):
        # 結合セルの範囲を探す
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # 親セル（左上）の座標を取得
                min_col = merged_range.min_col
                min_row = merged_range.min_row
                return ws.cell(row=min_row, column=min_col)
    return cell


def create_excel_minutes(data):
    """
    Excelファイルを作成し、議事録のテンプレートにデータを入力する。
    入力済みのExcelファイルを返す。
    """
    # ローカルファイルからテンプレートファイルを読み込む
    template_path = os.path.join(os.path.dirname(__file__), "../file/template.xlsx")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template file '{template_path}' not found.")
    with open(template_path, "rb") as f:
        template_excel = io.BytesIO(f.read())

    output_wb = openpyxl.load_workbook(template_excel)

    # Excelファイルを開く
    ws = output_wb["議事録_TEMPLATE"]
    # データを入力
    ws["F3"] = data["customer"] if data["customer"] else ""
    ws["F4"] = data["subject"] if data["subject"] else ""
    ws["F6"] = data["place"] if data["place"] else ""
    ws["F7"] = data["date_time"] if data["date_time"] else ""
    
    # 各セクションの構造（初期位置 - 行挿入により動的に変化）
    # 参加者: 8行目（ラベル）、9-13行目（データ）
    participants_label_row_initial = 8
    participants_start = 9
    participants_available = 5  # 9-13行目
    
    # 決定事項: 13行目（ラベル・初期位置）、14-19行目（データ）
    decisions_label_row_initial = 13
    decisions_start = 14
    decisions_available = 6  # 14-19行目
    
    # 宿題事項: 20行目（ラベル・初期位置）、21-25行目（データ）
    homework_label_row_initial = 20
    homework_start = 21
    homework_available = 5  # 21-25行目
    
    # 次回予定: 26行目（ラベル・初期位置）、27-28行目（データ）
    next_meeting_label_row_initial = 26
    next_meeting_row = 27
    
    # 会議内容: 29行目（ラベル・初期位置）、30行目以降（データ）
    minutes_label_row_initial = 29
    minutes_start = 30
    
    # 行挿入の累積数を追跡
    total_inserted = 0
    
    # 参加者の入力
    line = participants_start
    participants_count = sum(1 for item in data["participants"] for _ in item.keys())
    if participants_count > participants_available:
        insert_count = participants_count - participants_available
        # 参加者セクションの11行目をテンプレートとして使用
        template_row = 11  # 11行目
        # 決定事項データ開始行の前（decisions_startの前）に行を挿入（ラベル保護）
        insert_position = decisions_start  # 14行目の前
        # 空白行を1行追加（ラベル保護のため）
        ws.insert_rows(insert_position, insert_count + 1)
        # 挿入された行にフォーマットをコピー（最後の1行は空白行として残す）
        for i in range(insert_count):
            new_row = insert_position + i  # 挿入された実際の行
            copy_row_format(ws, template_row, new_row)
        total_inserted += insert_count + 1
    
    for item in data["participants"]:
        for company in item.keys():
            members = item[company]
            ws[f"F{line}"] = f"{company} : {', '.join(members)}"
            line += 1
    
    # 決定事項の入力（行挿入を考慮）
    line = decisions_start + total_inserted
    decisions_count = len(data["decision"])
    if decisions_count > decisions_available:
        insert_count = decisions_count - decisions_available
        # 18行目をテンプレートとして使用（最後の直前の行）
        template_row = 18 + total_inserted
        # 宿題事項ラベルの現在位置の前に行を挿入（初期20行目 + 累積挿入数）
        insert_position = homework_label_row_initial + total_inserted
        ws.insert_rows(insert_position, insert_count)
        # 挿入された行に18行目のフォーマットと結合セルをコピー
        for i in range(insert_count):
            new_row = insert_position + i
            copy_row_format(ws, template_row, new_row)
            copy_merged_cells(ws, template_row, new_row)
        total_inserted += insert_count
        # 行挿入後、移動した宿題事項ラベルの結合セルを再設定
        homework_label_current = homework_label_row_initial + total_inserted
        ensure_label_row_merged_cells(ws, homework_label_current, 'homework')
    
    for decision in data["decision"]:
        ws[f"C{line}"] = decision
        line += 1
    
    # 宿題事項の入力（行挿入を考慮）
    line = homework_start + total_inserted
    homework_count = len(data["homework"])
    if homework_count > homework_available:
        insert_count = homework_count - homework_available
        # 24行目をテンプレートとして使用（最後の直前の行）
        template_row = 24 + total_inserted
        # 次回予定ラベルの現在位置の前に行を挿入（初期26行目 + 累積挿入数）
        insert_position = next_meeting_label_row_initial + total_inserted
        ws.insert_rows(insert_position, insert_count)
        # 挿入された行に24行目のフォーマットと結合セルをコピー（C、AE、AG列すべて）
        for i in range(insert_count):
            new_row = insert_position + i
            copy_row_format(ws, template_row, new_row)
            copy_merged_cells(ws, template_row, new_row)
        total_inserted += insert_count
        # 行挿入後、移動した次回予定ラベルの結合セルを再設定
        next_meeting_label_current = next_meeting_label_row_initial + total_inserted
        ensure_label_row_merged_cells(ws, next_meeting_label_current, 'next_meeting')
    
    for item in data["homework"]:
        # トピック
        cell_c = get_merged_cell_parent(ws, ws[f"C{line}"])
        cell_c.value = item["topic"]
        # 担当者
        cell_ae = get_merged_cell_parent(ws, ws[f"AE{line}"])
        cell_ae.value = (
            item["item"]["responsible"] if item["item"]["responsible"] else "不明"
        )
        cell_ae.alignment = Alignment(horizontal="left")
        # 期限
        cell_ag = get_merged_cell_parent(ws, ws[f"AG{line}"])
        cell_ag.value = (
            item["item"]["deadline"] if item["item"]["deadline"] else "なし"
        )
        cell_ag.alignment = Alignment(horizontal="left")
        line += 1
    
    # 次回会議の日時の入力（行挿入を考慮）
    # 次回会議は通常1-2行で収まるが、必要に応じて行挿入可能
    next_meeting_available = 2  # 27-28行目（2行分）
    next_meeting_data = [data["next_meeting"]] if data["next_meeting"] else []
    next_meeting_count = len(next_meeting_data)
    
    if next_meeting_count > next_meeting_available:
        insert_count = next_meeting_count - next_meeting_available
        # 28行目をテンプレートとして使用
        template_row = 28 + total_inserted
        # 会議内容ラベルの現在位置の前に行を挿入（初期29行目 + 累積挿入数）
        insert_position = minutes_label_row_initial + total_inserted
        ws.insert_rows(insert_position, insert_count)
        # 挿入された行に28行目のフォーマットと結合セルをコピー
        for i in range(insert_count):
            new_row = insert_position + i
            copy_row_format(ws, template_row, new_row)
            copy_merged_cells(ws, template_row, new_row)
        total_inserted += insert_count
    
    # 会議内容ラベル（B29からAH29）の結合セルを保護（行挿入後の位置で再設定）
    minutes_label_current = minutes_label_row_initial + total_inserted
    ensure_label_row_merged_cells(ws, minutes_label_current, 'minutes')
    
    next_meeting_current_row = next_meeting_row + total_inserted
    ws[f"C{next_meeting_current_row}"] = data["next_meeting"] if data["next_meeting"] else ""
    
    # 議事録の内容入力（行挿入を考慮）
    line = minutes_start + total_inserted
    for minute in data["minutes"]:
        for topic, items in minute.items():
            # トピック（太字）
            cell_topic = get_merged_cell_parent(ws, ws[f"C{line}"])
            cell_topic.value = topic
            cell_topic.font = Font(bold=True)
            line += 1
            # 各内容
            for content in items:
                cell_content = get_merged_cell_parent(ws, ws[f"C{line}"])
                cell_content.value = content
                # セルのテキストを折り返す、上揃えにする
                cell_content.alignment = Alignment(wrap_text=True, vertical="top")
                value = cell_content.value
                max_length = 70  # 1行あたりの最大文字数
                if value:
                    lines = value.count("\n") + 1
                    est_lines = max(lines, math.ceil(len(value) / max_length))
                    # 12は1行分の高さ
                    ws.row_dimensions[cell_content.row].height = est_lines * 12
                line += 1
            line += 1
    # Excelファイルを保存
    output = io.BytesIO()
    output_wb.save(output)
    output.seek(0)

    # デバッグ用にファイルを保存する場合はコメントアウトを外す
    # output_wb.save("output.xlsx")
    return output
